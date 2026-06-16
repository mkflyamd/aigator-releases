"""Excel skill -- 6 tools."""

import subprocess
from pathlib import Path

from proc_utils import no_window_kwargs

SCRIPTS_DIR = Path(__file__).parent / "scripts"
SHARED_SCRIPTS = Path(__file__).parent.parent / "_scripts"
RECALC_SCRIPT = SCRIPTS_DIR / "recalc.py"

SKILL_ID = "excel"
SKILL_ALIASES = ["excel_skill"]

TOOL_DEFS = [
    {
        "name": "read_excel",
        "description": "Read cell values from an Excel workbook. Use BEFORE writing to understand the current data, headers, and layout. Use file_path='open' to read the currently open workbook via COM, or provide a full file path. Returns values as a 2D array with row/column info.",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Full path to .xlsx file, or 'open' for the active workbook via COM, or 'open:Budget.xlsx' to target a specific open workbook by name"},
                "sheet_name": {"type": "string", "description": "Worksheet name. Defaults to active sheet if omitted.", "default": ""},
                "cell": {"type": "string", "description": "Cell reference e.g. 'B3', or range 'A1:C10', or 'all' to read the entire used range"},
            },
            "required": ["file_path", "cell"],
        },
    },
    {
        "name": "list_excel_sheets",
        "description": "List all worksheet names in an Excel workbook. Use this first to discover what sheets exist before reading or writing. Use file_path='open' for the currently open workbook.",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Full path to .xlsx file, or 'open' for the active workbook via COM, or 'open:Budget.xlsx' to target a specific open workbook by name"},
            },
            "required": ["file_path"],
        },
    },
    {
        "name": "get_excel_info",
        "description": "Get structural info about an Excel worksheet: used range dimensions, column headers (first row), and total row/column count. Use this to understand layout before filling data \u2014 it tells you how many rows need filling without reading all data.",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Full path to .xlsx file, or 'open' for the active workbook via COM, or 'open:Budget.xlsx' to target a specific open workbook by name"},
                "sheet_name": {"type": "string", "description": "Worksheet name. Defaults to active sheet if omitted.", "default": ""},
            },
            "required": ["file_path"],
        },
    },
    {
        "name": "update_excel",
        "description": "Write values to cells. Supports single cell/row updates or batch mode for multiple rows in one call. ALWAYS use batch mode when writing more than one row. CRITICAL RULES: (1) Each column gets its own cell. (2) For a row range like 'A2:D2', use tab-separated values: 'Apple\\t5\\t3.99'. (3) Always write headers in row 1 first.",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Full path to .xlsx file, or 'open' for the active workbook via COM, or 'open:Budget.xlsx' to target a specific open workbook by name"},
                "sheet_name": {"type": "string", "description": "Worksheet name. Defaults to the first/active sheet if omitted.", "default": ""},
                "cell": {"type": "string", "description": "Single cell e.g. 'B3', or a single-row range e.g. 'A2:D2'. Not needed for batch mode."},
                "value": {"type": "string", "description": "For a single cell: the value. For a row range: tab-separated values. Not needed for batch mode."},
                "batch": {
                    "type": "array",
                    "description": "Array of row updates. Each has 'cell' (range like 'A2:D2') and 'value' (tab-separated). Use this to write multiple rows in one call.",
                    "items": {
                        "type": "object",
                        "properties": {
                            "cell": {"type": "string", "description": "Cell or range e.g. 'A2:D2'"},
                            "value": {"type": "string", "description": "Value or tab-separated values"},
                        },
                        "required": ["cell", "value"],
                    },
                },
            },
            "required": ["file_path"],
        },
    },
    {
        "name": "create_excel",
        "description": "Create a professional Excel workbook with formatted sheets, headers, data, and formulas. Supports multiple sheets in one call. Headers get bold text with shading, auto-filter, and freeze panes. Formulas (values starting with '=') are preserved. Call ONCE with all data.",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Full path where the .xlsx file will be saved"},
                "sheets": {
                    "type": "array",
                    "description": "Array of sheet definitions. Each sheet has name, headers, rows, and optional formatting.",
                    "items": {
                        "type": "object",
                        "properties": {
                            "name": {"type": "string", "description": "Worksheet name"},
                            "headers": {"type": "array", "items": {"type": "string"}, "description": "Column header names"},
                            "rows": {"type": "array", "items": {"type": "array", "items": {"type": "string"}}, "description": "Data rows as 2D array. Use '=' prefix for formulas, e.g. '=SUM(B2:B10)'"},
                            "column_widths": {"type": "array", "items": {"type": "number"}, "description": "Column widths in characters. Optional — auto-fit if omitted."},
                            "freeze_panes": {"type": "string", "description": "Cell ref to freeze at, e.g. 'A2' freezes header row. Defaults to 'A2'.", "default": "A2"},
                        },
                        "required": ["name", "headers", "rows"],
                    },
                },
                "author": {"type": "string", "description": "Workbook author metadata. Optional.", "default": ""},
            },
            "required": ["file_path", "sheets"],
        },
    },
    {
        "name": "recalc_excel",
        "description": "Recalculate all formulas in an Excel file and scan for errors (#REF!, #DIV/0!, #VALUE!, #N/A, #NAME?, #NULL!, #NUM!). Use after creating or editing workbooks with formulas. Requires LibreOffice installed.",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Full path to the .xlsx file to recalculate"},
                "timeout": {"type": "integer", "description": "Timeout in seconds. Defaults to 30.", "default": 30},
            },
            "required": ["file_path"],
        },
    },
]

TOOL_STATUS = {
    "read_excel": "\U0001f4ca Reading Excel...",
    "list_excel_sheets": "\U0001f4ca Listing Excel sheets...",
    "get_excel_info": "\U0001f4ca Inspecting Excel layout...",
    "update_excel": "\U0001f4c8 Updating Excel...",
    "create_excel": "\U0001f4ca Creating Excel workbook...",
    "recalc_excel": "\U0001f4ca Recalculating formulas...",
}


def _tool_read_excel(file_path: str, cell: str, sheet_name: str = "") -> dict:
    try:
        if file_path.startswith("open"):
            from skills._office_com import get_excel_app, get_excel_workbook
            app, err = get_excel_app()
            if err:
                return {"error": err}
            wb, err = get_excel_workbook(app, file_path)
            if err:
                return {"error": err}
            ws = wb.Sheets(sheet_name) if sheet_name else wb.ActiveSheet
            if cell.lower() == "all":
                used = ws.UsedRange
                data = used.Value
            else:
                data = ws.Range(cell).Value
            # Normalize to 2D list
            if data is None:
                rows = []
            elif not isinstance(data, tuple):
                rows = [[data]]
            else:
                rows = [list(r) if isinstance(r, tuple) else [r] for r in data]
            return {"ok": True, "workbook": wb.Name, "sheet": ws.Name, "cell": cell, "values": rows}
        else:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            if cell.lower() == "all":
                rows = [[c.value for c in row] for row in ws.iter_rows()]
            elif ":" in cell:
                rows = [[c.value for c in row] for row in ws[cell]]
            else:
                rows = [[ws[cell].value]]
            return {"ok": True, "sheet": ws.title, "cell": cell, "values": rows}
    except Exception as e:
        return {"error": str(e)}


def _tool_list_excel_sheets(file_path: str) -> dict:
    try:
        if file_path.startswith("open"):
            from skills._office_com import get_excel_app, get_excel_workbook, list_excel_workbooks
            app, err = get_excel_app()
            if err:
                return {"error": err}
            wb, err = get_excel_workbook(app, file_path)
            if err:
                return {"error": err}
            sheets = [wb.Sheets(i + 1).Name for i in range(wb.Sheets.Count)]
            result = {"ok": True, "workbook": wb.Name, "sheets": sheets, "active_sheet": wb.ActiveSheet.Name}
            # Show all open workbooks so user knows what's available
            all_books = list_excel_workbooks(app)
            if len(all_books) > 1:
                result["all_open_workbooks"] = all_books
                result["hint"] = "Multiple workbooks open. Use file_path='open:filename.xlsx' to target a specific one."
            return result
        else:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            return {"ok": True, "sheets": wb.sheetnames, "active_sheet": wb.sheetnames[0]}
    except Exception as e:
        return {"error": str(e)}


def _tool_get_excel_info(file_path: str, sheet_name: str = "") -> dict:
    try:
        if file_path.startswith("open"):
            from skills._office_com import get_excel_app, get_excel_workbook
            app, err = get_excel_app()
            if err:
                return {"error": err}
            wb, err = get_excel_workbook(app, file_path)
            if err:
                return {"error": err}
            ws = wb.Sheets(sheet_name) if sheet_name else wb.ActiveSheet
            used = ws.UsedRange
            row_count = used.Rows.Count
            col_count = used.Columns.Count
            first_row = ws.Range(f"A1:{chr(64 + col_count)}1").Value
            headers = list(first_row[0]) if isinstance(first_row, tuple) else [first_row]
            return {
                "ok": True, "workbook": wb.Name, "sheet": ws.Name,
                "used_range": used.Address,
                "row_count": row_count, "col_count": col_count,
                "headers": headers,
                "data_rows": row_count - 1,
            }
        else:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            row_count = ws.max_row or 0
            col_count = ws.max_column or 0
            return {
                "ok": True, "sheet": ws.title,
                "row_count": row_count, "col_count": col_count,
                "headers": headers,
                "data_rows": row_count - 1,
            }
    except Exception as e:
        return {"error": str(e)}


def _tool_update_excel(file_path: str, cell: str = "", value: str = "",
                        sheet_name: str = "", batch: list = None) -> dict:
    try:
        if batch:
            # Batch mode — multiple cell/row updates in one call
            if file_path.startswith("open"):
                from skills._office_com import get_excel_app, get_excel_workbook, save_com_document, get_file_info
                app, err = get_excel_app()
                if err:
                    return {"error": err}
                wb, err = get_excel_workbook(app, file_path)
                if err:
                    return {"error": err}
                finfo = get_file_info(wb, "excel")
                ws = wb.Sheets(sheet_name) if sheet_name else wb.ActiveSheet
                for op in batch:
                    ws.Range(op["cell"]).Value = op["value"]
                save_com_document(wb, "excel")
                return {"ok": True, **finfo, "message": f"Batch: updated {len(batch)} cells/rows in {finfo['file_name']}"}
            else:
                import openpyxl
                wb = openpyxl.load_workbook(file_path)
                ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
                for op in batch:
                    _write_cell(ws, op["cell"], op["value"])
                wb.save(file_path)
                return {"ok": True, "message": f"Batch: updated {len(batch)} cells/rows in {file_path}"}

        if not cell or value is None:
            return {"error": "cell and value are required (or use batch array)"}

        if file_path.startswith("open"):
            from skills._office_com import get_excel_app, get_excel_workbook, save_com_document, get_file_info
            app, err = get_excel_app()
            if err:
                return {"error": err}
            wb, err = get_excel_workbook(app, file_path)
            if err:
                return {"error": err}
            finfo = get_file_info(wb, "excel")
            ws = wb.Sheets(sheet_name) if sheet_name else wb.ActiveSheet
            ws.Range(cell).Value = value
            save_com_document(wb, "excel")
            return {"ok": True, **finfo, "message": f"Updated {cell} in {finfo['file_name']}"}
        else:
            import openpyxl
            wb = openpyxl.load_workbook(file_path)
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            _write_cell(ws, cell, value)
            wb.save(file_path)
            return {"ok": True, "message": f"Updated {cell} in {file_path}"}
    except Exception as e:
        return {"error": str(e)}


def _write_cell(ws, cell: str, value: str):
    """Write a value to a cell or range in an openpyxl worksheet."""
    if ":" in cell:
        start_cell = ws[cell.split(":")[0]]
        rows = [r.split("\t") for r in value.split("\n")]
        for r_idx, row in enumerate(rows):
            for c_idx, val in enumerate(row):
                ws.cell(row=start_cell.row + r_idx, column=start_cell.column + c_idx, value=val)
    else:
        ws[cell] = value


def _tool_create_excel(file_path: str, sheets: list, author: str = "") -> dict:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        # Remove default sheet
        if wb.sheetnames:
            wb.remove(wb.active)

        header_font = Font(name="Arial", size=11, bold=True)
        header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        normal_font = Font(name="Arial", size=11)
        thin_border = Border(
            left=Side(style="thin", color="BFBFBF"),
            right=Side(style="thin", color="BFBFBF"),
            top=Side(style="thin", color="BFBFBF"),
            bottom=Side(style="thin", color="BFBFBF"),
        )

        total_rows = 0
        total_sheets = 0
        has_formulas = False

        for sheet_def in sheets:
            ws = wb.create_sheet(title=sheet_def["name"])
            headers = sheet_def.get("headers", [])
            rows = sheet_def.get("rows", [])
            col_widths = sheet_def.get("column_widths")
            freeze = sheet_def.get("freeze_panes", "A2")

            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

            # Write data rows
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, val in enumerate(row_data, 1):
                    if isinstance(val, str) and val.startswith("="):
                        cell = ws.cell(row=row_idx, column=col_idx, value=val)
                        has_formulas = True
                    else:
                        # Try to convert numeric strings
                        try:
                            numeric = float(val)
                            if numeric == int(numeric) and "." not in str(val):
                                numeric = int(numeric)
                            cell = ws.cell(row=row_idx, column=col_idx, value=numeric)
                            cell.alignment = Alignment(horizontal="right")
                        except (ValueError, TypeError):
                            cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = normal_font
                    cell.border = thin_border

            # Column widths
            if col_widths:
                for col_idx, width in enumerate(col_widths, 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = width
            else:
                # Auto-fit based on content
                for col_idx in range(1, len(headers) + 1):
                    max_len = len(str(headers[col_idx - 1])) if col_idx <= len(headers) else 8
                    for row in ws.iter_rows(min_row=2, max_row=min(len(rows) + 1, 50), min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            if cell.value:
                                max_len = max(max_len, len(str(cell.value)))
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

            # Auto-filter on headers
            if headers:
                ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

            # Freeze panes
            if freeze:
                ws.freeze_panes = freeze

            total_rows += len(rows)
            total_sheets += 1

        if author:
            wb.properties.creator = author

        wb.save(file_path)

        # Attempt formula recalculation if formulas present
        recalc_msg = ""
        if has_formulas and RECALC_SCRIPT.exists():
            r = _tool_recalc_excel(file_path)
            if r.get("error"):
                recalc_msg = f" (recalc skipped: {r['error']})"
            elif r.get("error_count", 0) > 0:
                recalc_msg = f" (WARNING: {r['error_count']} formula errors found)"
            else:
                recalc_msg = " (formulas recalculated)"

        return {
            "ok": True,
            "message": f"Created {file_path} — {total_sheets} sheet(s), {total_rows} data rows{recalc_msg}",
            "file_path": file_path,
            "sheet_count": total_sheets,
            "row_count": total_rows,
            "has_formulas": has_formulas,
        }
    except Exception as e:
        return {"error": str(e)}


def _tool_recalc_excel(file_path: str, timeout: int = 30) -> dict:
    try:
        if not RECALC_SCRIPT.exists():
            return {"error": "recalc.py script not found. LibreOffice integration not available."}

        result = subprocess.run(
            ["python", str(RECALC_SCRIPT), str(file_path), str(timeout)],
            capture_output=True, text=True, timeout=timeout + 10,
            cwd=str(SHARED_SCRIPTS),
            **no_window_kwargs(),
        )

        output = result.stdout.strip()

        # Try to parse JSON output from recalc.py
        try:
            import json
            data = json.loads(output)
            return {
                "ok": True,
                "status": data.get("status", "unknown"),
                "total_formulas": data.get("total_formulas", 0),
                "error_count": data.get("total_errors", 0),
                "error_summary": data.get("error_summary", {}),
            }
        except (json.JSONDecodeError, ValueError):
            # Non-JSON output — return as message
            if result.returncode == 0:
                return {"ok": True, "message": output or "Recalculation completed"}
            else:
                return {"error": output or result.stderr.strip() or "Recalculation failed"}

    except subprocess.TimeoutExpired:
        return {"error": f"Recalculation timed out after {timeout}s"}
    except FileNotFoundError:
        return {"error": "LibreOffice not found. Install LibreOffice for formula recalculation."}
    except Exception as e:
        return {"error": str(e)}


TOOL_HANDLERS = {
    "read_excel": _tool_read_excel,
    "list_excel_sheets": _tool_list_excel_sheets,
    "get_excel_info": _tool_get_excel_info,
    "update_excel": _tool_update_excel,
    "create_excel": _tool_create_excel,
    "recalc_excel": _tool_recalc_excel,
}
